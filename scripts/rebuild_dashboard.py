"""
Regeneriert das KI-Use-Cases-Dashboard (HTML) aus der Excel-Datei.

Aufruf:
    python3 scripts/rebuild_dashboard.py                       # nutzt KI_UseCases_Uebersicht.xlsx im Repo-Root
    python3 scripts/rebuild_dashboard.py /pfad/zur/datei.xlsx  # expliziter Pfad

Liest das Tabellenblatt "Use Cases" und schreibt KI_UseCases_Dashboard.html
in denselben Ordner wie die Excel-Datei.
"""

import json
import os
import sys
from openpyxl import load_workbook


HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="de">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>KI Use Cases – Discovery Bundle Dashboard</title>
<style>
  :root {
    color-scheme: light;
    --bg: #f5f7fb;
    --card: #ffffff;
    --primary: #1f3a5f;
    --primary-light: #2d5a8a;
    --accent: #4a90e2;
    --text: #1a2238;
    --text-muted: #5b6478;
    --border: #e3e8f0;
    --tag-bg: #eef3fb;
    --tag-text: #1f3a5f;
    --shadow: 0 2px 8px rgba(31, 58, 95, 0.06);
    --shadow-hover: 0 6px 20px rgba(31, 58, 95, 0.12);
  }
  * { box-sizing: border-box; }
  body { margin: 0; font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Helvetica, Arial, sans-serif; background: var(--bg); color: var(--text); line-height: 1.5; }
  header { background: linear-gradient(135deg, var(--primary) 0%, var(--primary-light) 100%); color: white; padding: 28px 32px; box-shadow: var(--shadow); }
  .header-eyebrow { font-size: 12px; text-transform: uppercase; letter-spacing: 1px; opacity: 0.7; margin-bottom: 6px; }
  header h1 { margin: 0 0 4px 0; font-size: 26px; font-weight: 600; }
  header p { margin: 0; opacity: 0.85; font-size: 14px; }
  .container { max-width: 1400px; margin: 0 auto; padding: 24px 32px 60px; }
  .stats { display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr)); gap: 16px; margin-bottom: 24px; }
  .stat { background: var(--card); border-radius: 10px; padding: 18px 20px; box-shadow: var(--shadow); border-left: 4px solid var(--accent); }
  .stat-label { font-size: 12px; color: var(--text-muted); text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 6px; }
  .stat-value { font-size: 28px; font-weight: 700; color: var(--primary); }
  .filters { background: var(--card); border-radius: 12px; padding: 22px 24px; box-shadow: var(--shadow); margin-bottom: 24px; }
  .filter-row { display: grid; grid-template-columns: 1fr 1fr; gap: 24px; margin-bottom: 16px; }
  @media (max-width: 800px) { .filter-row { grid-template-columns: 1fr; } }
  .filter-group h3 { margin: 0 0 10px 0; font-size: 13px; color: var(--text-muted); text-transform: uppercase; letter-spacing: 0.5px; font-weight: 600; }
  .chips { display: flex; flex-wrap: wrap; gap: 8px; }
  .chip { background: var(--tag-bg); color: var(--tag-text); padding: 7px 14px; border-radius: 16px; font-size: 13px; cursor: pointer; border: 1.5px solid transparent; transition: all 0.15s; user-select: none; display: inline-flex; align-items: center; gap: 6px; }
  .chip:hover { background: #dde6f3; }
  .chip.active { background: var(--primary); color: white; border-color: var(--primary); }
  .chip .count { font-size: 11px; opacity: 0.7; background: rgba(0,0,0,0.05); padding: 1px 6px; border-radius: 8px; }
  .chip.active .count { background: rgba(255,255,255,0.2); opacity: 1; }
  .search-and-actions { display: flex; gap: 12px; align-items: center; border-top: 1px solid var(--border); padding-top: 16px; }
  .search { flex: 1; padding: 10px 14px; border: 1px solid var(--border); border-radius: 8px; font-size: 14px; font-family: inherit; outline: none; transition: border-color 0.15s; }
  .search:focus { border-color: var(--accent); }
  .btn-reset { background: none; border: 1px solid var(--border); color: var(--text-muted); padding: 10px 18px; border-radius: 8px; font-size: 13px; cursor: pointer; transition: all 0.15s; font-family: inherit; }
  .btn-reset:hover { border-color: var(--primary); color: var(--primary); }
  .result-info { margin-bottom: 16px; color: var(--text-muted); font-size: 14px; }
  .result-info strong { color: var(--primary); }
  .cards { display: grid; grid-template-columns: repeat(auto-fill, minmax(380px, 1fr)); gap: 16px; }
  .card { background: var(--card); border-radius: 12px; padding: 22px; box-shadow: var(--shadow); transition: all 0.2s; border: 1px solid var(--border); display: flex; flex-direction: column; }
  .card:hover { box-shadow: var(--shadow-hover); transform: translateY(-2px); }
  .card-tags { display: flex; gap: 6px; flex-wrap: wrap; margin-bottom: 12px; }
  .tag { font-size: 11px; padding: 3px 10px; border-radius: 10px; font-weight: 500; text-transform: uppercase; letter-spacing: 0.3px; }
  .tag-abt { background: #eef3fb; color: #1f3a5f; }
  .tag-art { background: #f0f7ee; color: #2d6a3e; }
  .card-section { margin-bottom: 12px; }
  .card-section-label { font-size: 11px; color: var(--text-muted); text-transform: uppercase; letter-spacing: 0.5px; font-weight: 600; margin-bottom: 4px; }
  .card-problem { font-size: 14px; color: var(--text); line-height: 1.5; }
  .card-solution { font-size: 15px; font-weight: 600; color: var(--primary); line-height: 1.4; }
  .empty { grid-column: 1 / -1; text-align: center; padding: 60px 20px; color: var(--text-muted); background: var(--card); border-radius: 12px; }
  .empty-icon { font-size: 40px; margin-bottom: 12px; opacity: 0.4; }
  footer { text-align: center; color: var(--text-muted); font-size: 12px; margin-top: 40px; padding: 20px; }
</style>
</head>
<body>

<header>
  <div class="header-eyebrow">Discovery Bundle · Business Workshop</div>
  <h1>KI Use Cases aus dem Discovery Bundle</h1>
  <p>Uebersicht der im Business Workshop identifizierten KI-Anwendungsfaelle. Filtere nach Abteilung und Use Case Art, um relevante Use Cases zu entdecken.</p>
</header>

<div class="container">

  <div class="stats">
    <div class="stat"><div class="stat-label">Use Cases gesamt</div><div class="stat-value" id="stat-total">-</div></div>
    <div class="stat"><div class="stat-label">Abteilungen</div><div class="stat-value" id="stat-abt">-</div></div>
    <div class="stat"><div class="stat-label">Use Case Arten</div><div class="stat-value" id="stat-art">-</div></div>
    <div class="stat"><div class="stat-label">Aktuell angezeigt</div><div class="stat-value" id="stat-shown">-</div></div>
  </div>

  <div class="filters">
    <div class="filter-row">
      <div class="filter-group">
        <h3>Abteilung</h3>
        <div class="chips" id="chips-abt"></div>
      </div>
      <div class="filter-group">
        <h3>Use Case Art</h3>
        <div class="chips" id="chips-art"></div>
      </div>
    </div>
    <div class="search-and-actions">
      <input type="text" class="search" id="search" placeholder="In Problem oder KI-Loesung suchen...">
      <button class="btn-reset" id="reset">Filter zuruecksetzen</button>
    </div>
  </div>

  <div class="result-info" id="result-info"></div>
  <div class="cards" id="cards"></div>

</div>

<footer>
  KI Use Cases &middot; Quelle: Discovery Bundle / Business Workshop &middot; Stand: <span id="updated"></span>
</footer>

<script>
const USE_CASES = __DATA_PLACEHOLDER__;

const state = { abt: new Set(), art: new Set(), search: "" };

function countBy(items, key) {
  const c = {};
  items.forEach(i => { c[i[key]] = (c[i[key]] || 0) + 1; });
  return c;
}
function uniqueSorted(items, key) {
  return [...new Set(items.map(i => i[key]))].sort((a,b) => a.localeCompare(b, 'de'));
}

function renderChips(containerId, key, stateSet) {
  const container = document.getElementById(containerId);
  const counts = countBy(USE_CASES, key);
  const values = uniqueSorted(USE_CASES, key);
  container.innerHTML = "";
  values.forEach(val => {
    const chip = document.createElement("div");
    chip.className = "chip" + (stateSet.has(val) ? " active" : "");
    chip.innerHTML = `${val} <span class="count">${counts[val]}</span>`;
    chip.addEventListener("click", () => {
      if (stateSet.has(val)) stateSet.delete(val); else stateSet.add(val);
      render();
    });
    container.appendChild(chip);
  });
}

function render() {
  const search = state.search.toLowerCase().trim();
  const filtered = USE_CASES.filter(uc => {
    if (state.abt.size && !state.abt.has(uc.abt)) return false;
    if (state.art.size && !state.art.has(uc.art)) return false;
    if (search) {
      const hay = (uc.p + " " + uc.l).toLowerCase();
      if (!hay.includes(search)) return false;
    }
    return true;
  });

  document.getElementById("stat-shown").textContent = filtered.length;
  const info = document.getElementById("result-info");
  if (filtered.length === USE_CASES.length) {
    info.innerHTML = `Zeige <strong>alle ${filtered.length}</strong> Use Cases.`;
  } else {
    info.innerHTML = `Zeige <strong>${filtered.length}</strong> von ${USE_CASES.length} Use Cases.`;
  }

  const cards = document.getElementById("cards");
  cards.innerHTML = "";
  if (filtered.length === 0) {
    cards.innerHTML = `<div class="empty"><div class="empty-icon">&#128269;</div><div>Keine Use Cases gefunden. Passe Filter oder Suchbegriff an.</div></div>`;
    return;
  }
  filtered.forEach(uc => {
    const card = document.createElement("div");
    card.className = "card";
    card.innerHTML = `
      <div class="card-tags">
        <span class="tag tag-abt">${uc.abt}</span>
        <span class="tag tag-art">${uc.art}</span>
      </div>
      <div class="card-section">
        <div class="card-section-label">Problem</div>
        <div class="card-problem">${uc.p}</div>
      </div>
      <div class="card-section">
        <div class="card-section-label">KI-Loesung</div>
        <div class="card-solution">${uc.l}</div>
      </div>
    `;
    cards.appendChild(card);
  });

  renderChips("chips-abt", "abt", state.abt);
  renderChips("chips-art", "art", state.art);
}

function init() {
  document.getElementById("stat-total").textContent = USE_CASES.length;
  document.getElementById("stat-abt").textContent = uniqueSorted(USE_CASES, "abt").length;
  document.getElementById("stat-art").textContent = uniqueSorted(USE_CASES, "art").length;
  document.getElementById("search").addEventListener("input", e => { state.search = e.target.value; render(); });
  document.getElementById("reset").addEventListener("click", () => {
    state.abt.clear(); state.art.clear(); state.search = "";
    document.getElementById("search").value = "";
    render();
  });
  const d = new Date();
  document.getElementById("updated").textContent = d.toLocaleDateString("de-DE", {year:"numeric", month:"long", day:"numeric"});
  render();
}

init();
</script>
</body>
</html>
"""


def find_default_excel():
    """Sucht KI_UseCases_Uebersicht.xlsx im Repo-Root (Skript liegt in scripts/) oder im cwd."""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    candidates = [
        os.path.join(script_dir, "..", "KI_UseCases_Uebersicht.xlsx"),
        os.path.join(script_dir, "KI_UseCases_Uebersicht.xlsx"),
        os.path.join(os.getcwd(), "KI_UseCases_Uebersicht.xlsx"),
    ]
    for c in candidates:
        if os.path.isfile(c):
            return os.path.abspath(c)
    return None


def main():
    if len(sys.argv) >= 2:
        excel_path = sys.argv[1]
    else:
        excel_path = find_default_excel()
        if not excel_path:
            print("Fehler: KI_UseCases_Uebersicht.xlsx nicht gefunden. Pfad als Argument übergeben.")
            sys.exit(1)

    if not os.path.isfile(excel_path):
        print(f"Fehler: Datei nicht gefunden: {excel_path}")
        sys.exit(1)

    wb = load_workbook(excel_path, data_only=True)
    if "Use Cases" not in wb.sheetnames:
        print("Fehler: Tabellenblatt 'Use Cases' nicht gefunden.")
        sys.exit(1)

    ws = wb["Use Cases"]
    headers = [c.value for c in ws[1]]

    def col(name):
        if name not in headers:
            raise SystemExit(f"Fehler: Spalte '{name}' nicht gefunden. Header: {headers}")
        return headers.index(name) + 1

    p_idx = col("Problem")
    l_idx = col("KI-Lösung")
    abt_idx = col("Abteilung")
    art_idx = col("Use Case Art")
    quelle_idx = col("Quelle Workshop") if "Quelle Workshop" in headers else None

    data = []
    for r in range(2, ws.max_row + 1):
        p = ws.cell(row=r, column=p_idx).value
        l = ws.cell(row=r, column=l_idx).value
        if not p or not l:
            continue
        entry = {
            "p": str(p).strip(),
            "l": str(l).strip(),
            "abt": str(ws.cell(row=r, column=abt_idx).value or "").strip(),
            "art": str(ws.cell(row=r, column=art_idx).value or "").strip(),
        }
        if quelle_idx:
            entry["quelle"] = str(ws.cell(row=r, column=quelle_idx).value or "").strip()
        data.append(entry)

    if not data:
        print("Fehler: Keine Use Cases im Tabellenblatt 'Use Cases' gefunden.")
        sys.exit(1)

    data_json = json.dumps(data, ensure_ascii=False, indent=0)
    html = HTML_TEMPLATE.replace("__DATA_PLACEHOLDER__", data_json)

    out_path = os.path.join(os.path.dirname(os.path.abspath(excel_path)), "KI_UseCases_Dashboard.html")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"Dashboard geschrieben: {out_path}")
    print(f"Use Cases im Dashboard: {len(data)}")


if __name__ == "__main__":
    main()
